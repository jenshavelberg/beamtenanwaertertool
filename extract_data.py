import openpyxl, json

wb = openpyxl.load_workbook('/Users/jens/Documents/git/beamtenanwaertertool/Beamtenanwärter-Tool.xlsx', data_only=True)
ws = wb['Berechnung']

bundeslaender = [
    (4, 7, 20, "Baden-Württemberg"),
    (22, 25, 38, "Bayern"),
    (40, 43, 56, "Berlin"),
    (58, 61, 74, "Brandenburg"),
    (76, 79, 92, "Bremen"),
    (94, 97, 110, "Hamburg"),
    (112, 115, 128, "Hessen"),
    (130, 133, 146, "Mecklenburg-Vorpommern"),
    (148, 151, 164, "Niedersachsen"),
    (166, 169, 182, "Nordrhein-Westfalen"),
    (184, 187, 200, "Rheinland-Pfalz"),
    (202, 205, 218, "Saarland"),
    (220, 223, 236, "Sachsen"),
    (238, 241, 254, "Sachsen-Anhalt"),
    (256, 259, 272, "Schleswig-Holstein"),
    (274, 277, 290, "Thüringen"),
    (294, 297, 310, "Bundesbeihilfe"),
]

all_data = {}

for header_row, data_start, data_end, name in bundeslaender:
    entries = []
    for r in range(data_start, data_end + 1):
        b_val = ws.cell(row=r, column=2).value
        if not b_val or not str(b_val).strip():
            continue
        
        dienstgrad = str(b_val).strip()
        c_val = ws.cell(row=r, column=3).value
        d_val = ws.cell(row=r, column=4).value
        
        def safe_float(v):
            if v is None: return 0.0
            try: return float(v)
            except: return 0.0
        
        entry = {
            "dienstgrad": dienstgrad,
            "familienstand": str(c_val) if c_val else "",
            "besoldungsgruppe": str(d_val) if d_val else "",
            "grundbezuege": safe_float(ws.cell(row=r, column=5).value),
            "famZuschlag": safe_float(ws.cell(row=r, column=6).value),
            "vl": safe_float(ws.cell(row=r, column=7).value),
            "brutto_ledig": safe_float(ws.cell(row=r, column=8).value),
            "stkl_ledig": str(ws.cell(row=r, column=9).value or ""),
            "lst_ledig": safe_float(ws.cell(row=r, column=10).value),
            "soli_ledig": safe_float(ws.cell(row=r, column=11).value),
            "kist_ledig": safe_float(ws.cell(row=r, column=12).value),
            "steuer_ledig": safe_float(ws.cell(row=r, column=13).value),
            "netto_ledig": safe_float(ws.cell(row=r, column=14).value),
            "kvSatz_ledig": safe_float(ws.cell(row=r, column=15).value),
            "brutto_verh": safe_float(ws.cell(row=r, column=16).value),
            "stkl_verh": str(ws.cell(row=r, column=17).value or ""),
            "lst_verh": safe_float(ws.cell(row=r, column=18).value),
            "soli_verh": safe_float(ws.cell(row=r, column=19).value),
            "kist_verh": safe_float(ws.cell(row=r, column=20).value),
            "steuer_verh": safe_float(ws.cell(row=r, column=21).value),
            "netto_verh": safe_float(ws.cell(row=r, column=22).value),
            "kvSatz_verh": safe_float(ws.cell(row=r, column=23).value),
        }
        entries.append(entry)
    
    all_data[name] = entries
    print(f"{name}: {len(entries)} Eintraege")

with open('/Users/jens/Documents/git/beamtenanwaertertool/besoldungsdaten.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print(f"\nGesamt: {sum(len(v) for v in all_data.values())} Datensaetze")
